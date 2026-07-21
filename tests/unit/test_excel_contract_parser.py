import openpyxl
import pytest

from api_quality_agent.domain.exceptions import (
    CorruptedInputFileError,
    EmptyInputError,
    InputFileNotFoundError,
)
from api_quality_agent.domain.models import DeclaredSchema, ParameterLocation
from api_quality_agent.parsers import ExcelContractParser

_MODEL_SHEET_ROWS: list[list[object]] = [
    ["Situação", "campo disponível"],
    ["Operação", "NOVO"],
    ["Número da operação IIB", None],
    ["Número da versão da operação IIB", None],
    ["Nome do módulo no BB", None],
    ["URI", "/teste/v{version}/cotacao/{origem}/contexto/{context}/canal/{canal}/ageteteste/coberturas"],
    ["Método", "POST"],
    [],
    ["Requisição (HEADER)"],
    ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    [1, "nmIdTransacaoExterno", "Alfanumerico", 40, "SIM"],
    [],
    ["Requisição (Path Param)"],
    ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    [1, "origem", "Alfanumerico", 40, "SIM"],
    [],
    ["Requisição (Query Param)"],
    ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    [],
    ["Requisição(Body)"],
    ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    [1, "nrOferta", "Numérico", 10, "NÃO"],
    [2, "nrVersaoOferta", "Numérico", 1, "SIM"],
    [],
    ["Resposta caso HTTP Status code 200 - OK"],
    ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    ["1", "dado", "Objeto", None, "SIM"],
    ["1.1", "id", "Alfanumerico", 10, "SIM"],
    ["1.2", "tipo", "Alfanumerico", 50, "SIM"],
    ["1.3", "atributos", "Objeto", None, "SIM"],
    ["1.3.1", "lsObjetoSegurado", "ArrayList", 30, "SIM"],
    ["1.3.1.1", "idObjetoSegurado", "Numérico", 10, "SIM"],
    [],
    ["Resposta caso HTTP Status code 400 - ERRO"],
    ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    ["1", "erros", "Lista", 10, None],
    ["1.1", "codTipoMensagemErro", "Numérico", 5, None],
]


def _build_workbook(tmp_path, *, sheets: dict[str, list[list[object]]]):
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    first = True
    for title, rows in sheets.items():
        sheet = default_sheet if first else workbook.create_sheet()
        sheet.title = title
        for row in rows:
            sheet.append(row)
        first = False
    path = tmp_path / "contrato.xlsx"
    workbook.save(path)
    return path


def _find(properties: tuple[DeclaredSchema, ...], name: str) -> DeclaredSchema:
    return next(prop for prop in properties if prop.name == name)


def test_parses_metadata_method_and_uri(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    catalog = ExcelContractParser().parse(path).catalog

    assert len(catalog.contracts) == 1
    endpoint = catalog.contracts[0]
    assert endpoint.method == "POST"
    assert endpoint.path == "/teste/v{version}/cotacao/{origem}/contexto/{context}/canal/{canal}/ageteteste/coberturas"
    assert endpoint.source_sheet == "Planilha1"


def test_source_file_is_recorded(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    catalog = ExcelContractParser().parse(path).catalog

    assert catalog.source_file == str(path)


def test_header_becomes_a_required_string_parameter(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    assert len(endpoint.request.headers) == 1
    header = endpoint.request.headers[0]
    assert header.name == "nmIdTransacaoExterno"
    assert header.location is ParameterLocation.HEADER
    assert header.required is True
    assert header.schema.type == "string"


def test_path_param_becomes_a_required_string_parameter(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    assert len(endpoint.request.path_parameters) == 1
    path_parameter = endpoint.request.path_parameters[0]
    assert path_parameter.name == "origem"
    assert path_parameter.location is ParameterLocation.PATH
    assert path_parameter.required is True


def test_empty_query_param_section_produces_no_parameters(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    assert endpoint.request.query_parameters == ()


def test_body_schema_has_optional_and_required_top_level_fields(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    body = endpoint.request.body_schema
    assert body is not None
    assert body.type == "object"
    nr_oferta = _find(body.properties, "nrOferta")
    assert nr_oferta.type == "number"
    assert nr_oferta.required is False
    nr_versao = _find(body.properties, "nrVersaoOferta")
    assert nr_versao.required is True


def test_response_200_schema_builds_the_full_nested_tree(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    response_schema = endpoint.response.schema
    assert response_schema is not None
    dado = _find(response_schema.properties, "dado")
    assert dado.type == "object"
    assert dado.required is True

    atributos = _find(dado.properties, "atributos")
    assert atributos.type == "object"

    lista = _find(atributos.properties, "lsObjetoSegurado")
    assert lista.type == "array"
    assert lista.items is not None
    assert lista.items.type == "object"
    item_field = _find(lista.items.properties, "idObjetoSegurado")
    assert item_field.type == "number"
    assert item_field.required is True


def test_response_400_section_is_recognized_but_never_used(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    # Escopo R2-00B: só a resposta de sucesso (200) vira schema. A árvore de
    # "dado" (200) não deve ter nenhum resquício de "erros" (400).
    response_schema = endpoint.response.schema
    assert response_schema is not None
    names = {prop.name for prop in response_schema.properties}
    assert names == {"dado"}


def test_response_400_rows_still_appear_in_raw_rows(tmp_path):
    path = _build_workbook(tmp_path, sheets={"Planilha1": _MODEL_SHEET_ROWS})

    result = ExcelContractParser().parse(path)

    sections_present = {row.section for row in result.raw_rows}
    assert "response_400" in sections_present


def test_sheet_without_uri_or_metodo_produces_no_contract(tmp_path):
    rows_without_metadata = [
        ["Requisição(Body)"],
        ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
        [1, "campo", "Alfanumerico", 10, "SIM"],
    ]
    path = _build_workbook(tmp_path, sheets={"SemMetadados": rows_without_metadata})

    catalog = ExcelContractParser().parse(path).catalog

    assert catalog.contracts == ()


def test_multiple_sheets_each_become_a_separate_contract(tmp_path):
    other_sheet_rows = [
        ["URI", "/teste/v1/outra"],
        ["Método", "GET"],
        ["Requisição(Body)"],
        ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
    ]
    path = _build_workbook(
        tmp_path,
        sheets={"Planilha1": _MODEL_SHEET_ROWS, "Planilha2": other_sheet_rows},
    )

    catalog = ExcelContractParser().parse(path).catalog

    assert len(catalog.contracts) == 2
    methods = {contract.method for contract in catalog.contracts}
    assert methods == {"POST", "GET"}


def test_array_with_a_single_child_still_becomes_an_object_item_schema(tmp_path):
    rows = [
        ["URI", "/teste/v1/array-simples"],
        ["Método", "GET"],
        ["Resposta caso HTTP Status code 200 - OK"],
        ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"],
        ["1", "itens", "ArrayList", 5, "SIM"],
        ["1.1", "codigo", "Numérico", 5, "SIM"],
    ]
    path = _build_workbook(tmp_path, sheets={"Planilha1": rows})

    endpoint = ExcelContractParser().parse(path).catalog.contracts[0]

    itens = _find(endpoint.response.schema.properties, "itens")  # type: ignore[union-attr]
    assert itens.type == "array"
    assert itens.items is not None
    assert itens.items.type == "object"
    assert itens.items.properties[0].name == "codigo"


# Correção dos exit codes: arquivo de contrato inexistente/vazio/corrompido/
# sem permissão/caminho inválido deve virar um InputError (exit code 2), não
# escapar como exceção crua (o que hoje cairia no catch-all de main.py e
# resultaria no exit code 8).


def test_nonexistent_contract_file_raises_input_file_not_found_error(tmp_path):
    missing_path = tmp_path / "nao_existe.xlsx"

    with pytest.raises(InputFileNotFoundError):
        ExcelContractParser().parse(missing_path)


def test_directory_as_contract_file_raises_input_file_not_found_error(tmp_path):
    directory_path = tmp_path / "diretorio.xlsx"
    directory_path.mkdir()

    with pytest.raises(InputFileNotFoundError):
        ExcelContractParser().parse(directory_path)


def test_empty_contract_file_raises_empty_input_error(tmp_path):
    empty_path = tmp_path / "vazio.xlsx"
    empty_path.write_bytes(b"")

    with pytest.raises(EmptyInputError):
        ExcelContractParser().parse(empty_path)


def test_corrupted_contract_file_raises_corrupted_input_file_error(tmp_path):
    corrupted_path = tmp_path / "corrompido.xlsx"
    corrupted_path.write_bytes(b"isto nao e um arquivo zip valido")

    with pytest.raises(CorruptedInputFileError):
        ExcelContractParser().parse(corrupted_path)


def test_unreadable_contract_file_raises_corrupted_input_file_error(tmp_path, monkeypatch):
    # PermissionError é um OSError levantado pelo próprio openpyxl ao tentar
    # abrir um arquivo sem permissão de leitura — simulado aqui via
    # monkeypatch porque negar permissão de leitura de forma confiável e
    # portável (Windows/POSIX) num arquivo real de teste não é viável.
    path = tmp_path / "sem_permissao.xlsx"
    path.write_bytes(b"conteudo irrelevante, load_workbook e substituido abaixo")

    def _raise_permission_error(*args, **kwargs):
        raise PermissionError(f"Sem permissão para ler {path}")

    monkeypatch.setattr(openpyxl, "load_workbook", _raise_permission_error)

    with pytest.raises(CorruptedInputFileError):
        ExcelContractParser().parse(path)
