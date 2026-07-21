import json
from pathlib import Path

import openpyxl
import pytest

from api_quality_agent.adapters.filesystem import LocalArtifactRepository
from api_quality_agent.application.use_cases import GenerateTestsWithContractUseCase
from api_quality_agent.domain.exceptions import InputError, StrictContractMatchFailedError
from api_quality_agent.domain.services import (
    ApiAnalysisEngine,
    DiffEngine,
    ManagedBlockMerger,
    SchemaInferenceEngine,
    TestStrategyEngine,
)
from api_quality_agent.generators import PostmanTestGenerator
from api_quality_agent.parsers import ExcelContractParser, PostmanCollectionParser
from api_quality_agent.domain.models import InputOrigin, ResolvedInput

_HEADER_ROW = ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"]


def _build_contract_workbook(tmp_path, entries):
    workbook = openpyxl.Workbook()
    first = True
    for sheet_title, method, uri in entries:
        sheet = workbook.active if first else workbook.create_sheet()
        sheet.title = sheet_title
        first = False
        rows = [
            ["URI", uri],
            ["Método", method],
            ["Resposta caso HTTP Status code 200 - OK"],
            _HEADER_ROW,
            ["1", "dado", "Objeto", None, "SIM"],
            ["1.1", "id", "Alfanumerico", 10, "SIM"],
        ]
        for row in rows:
            sheet.append(row)
    path = tmp_path / "contrato.xlsx"
    workbook.save(path)
    return str(path)


def _parse_document(items):
    document = {
        "info": {
            "name": "Col",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "item": items,
    }
    resolved = ResolvedInput(
        origin=InputOrigin.FILE, content_type="json", name="c.json", content=json.dumps(document)
    )
    return PostmanCollectionParser().parse(resolved)


def _build_use_case(tmp_path):
    artifact_repository = LocalArtifactRepository(tmp_path / "artifacts")
    return GenerateTestsWithContractUseCase(
        ExcelContractParser(),
        ApiAnalysisEngine(),
        SchemaInferenceEngine(),
        TestStrategyEngine(),
        PostmanTestGenerator(),
        ManagedBlockMerger(),
        DiffEngine(),
        artifact_repository,
    )


def test_execute_offline_saves_a_contract_match_report_json_and_html(tmp_path):
    contract_path = _build_contract_workbook(
        tmp_path, [("Planilha1", "GET", "/v2/pet/{{petId}}")]
    )
    document = _parse_document(
        [
            {
                "name": "Buscar pet",
                "id": "r1",
                "request": {"method": "GET", "url": "https://x/v2/pet/:petId"},
                "response": [
                    {"name": "ok", "status": "OK", "code": 200, "header": [], "body": "{}"}
                ],
            }
        ]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=contract_path, document=document)

    json_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.json")
    )
    html_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.html")
    )

    payload = json.loads(Path(json_location.path).read_text(encoding="utf-8"))
    assert payload["summary"]["matched"] == 1
    assert payload["summary"]["not_found"] == 0
    assert payload["matches"][0]["status"] == "MATCHED"

    html_content = Path(html_location.path).read_text(encoding="utf-8")
    assert "MATCHED" in html_content


def test_execute_offline_reports_not_found_when_request_has_no_declared_contract(tmp_path):
    contract_path = _build_contract_workbook(tmp_path, [("Planilha1", "GET", "/v2/outra-coisa")])
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/1"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=contract_path, document=document)

    json_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(json_location.path).read_text(encoding="utf-8"))
    assert payload["summary"]["not_found"] == 1
    assert payload["summary"]["matched"] == 0


def test_match_report_is_saved_in_the_same_execution_folder_as_scripts_and_diff(tmp_path):
    contract_path = _build_contract_workbook(
        tmp_path, [("Planilha1", "GET", "/v2/pet/{{petId}}")]
    )
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/:petId"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=contract_path, document=document)

    directories = {Path(loc.path).parent.parent for loc in result.artifact_locations}
    assert len(directories) == 1  # todos os artefatos (scripts/diffs/contracts) na mesma pasta de execução


def test_execute_online_without_workspace_dependencies_raises_input_error(tmp_path):
    use_case = _build_use_case(tmp_path)

    with pytest.raises(InputError):
        use_case.execute_online(contract_file="qualquer.xlsx", collection_id="abc")


def test_execute_offline_without_collection_path_prefix_reports_not_found_when_collection_has_gateway_prefix(
    tmp_path,
):
    # Contrato declara o path sem o prefixo de gateway; a Collection real
    # usa o prefixo — sem a flag, não deve casar (comportamento atual).
    contract_path = _build_contract_workbook(tmp_path, [("Planilha1", "GET", "/v1/users/{id}")])
    document = _parse_document(
        [{"name": "Buscar usuário", "id": "r1", "request": {"method": "GET", "url": "https://x/api/v1/users/:id"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=contract_path, document=document)

    json_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(json_location.path).read_text(encoding="utf-8"))
    assert payload["summary"]["not_found"] == 1
    assert payload["summary"]["matched"] == 0


def test_execute_offline_with_collection_path_prefix_matches_the_declared_contract(tmp_path):
    contract_path = _build_contract_workbook(tmp_path, [("Planilha1", "GET", "/v1/users/{id}")])
    document = _parse_document(
        [{"name": "Buscar usuário", "id": "r1", "request": {"method": "GET", "url": "https://x/api/v1/users/:id"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(
        contract_file=contract_path, document=document, collection_path_prefix="/api"
    )

    json_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(json_location.path).read_text(encoding="utf-8"))
    assert payload["summary"]["matched"] == 1
    assert payload["summary"]["not_found"] == 0


# --strict-contract-match: processa todos os endpoints e persiste o
# Contract Match Report normalmente; só depois decide falhar o comando se
# houver UNMATCHED/AMBIGUOUS.


def test_strict_contract_match_does_not_raise_when_everything_is_matched(tmp_path):
    contract_path = _build_contract_workbook(
        tmp_path, [("Planilha1", "GET", "/v2/pet/{{petId}}")]
    )
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/:petId"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(
        contract_file=contract_path, document=document, strict_contract_match=True
    )

    assert result.endpoint_outcomes[0].error is None


def test_strict_contract_match_raises_when_an_endpoint_is_not_found(tmp_path):
    contract_path = _build_contract_workbook(tmp_path, [("Planilha1", "GET", "/v2/outra-coisa")])
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/1"}}]
    )

    use_case = _build_use_case(tmp_path)

    with pytest.raises(StrictContractMatchFailedError) as exc_info:
        use_case.execute_offline(
            contract_file=contract_path, document=document, strict_contract_match=True
        )

    message = str(exc_info.value)
    assert "Unmatched: 1" in message
    assert "Ambiguous: 0" in message

    # O relatório precisa ter sido persistido ANTES da falha estrita.
    json_reports = list((tmp_path / "artifacts").rglob("contract-match-report.json"))
    html_reports = list((tmp_path / "artifacts").rglob("contract-match-report.html"))
    assert len(json_reports) == 1
    assert len(html_reports) == 1
    payload = json.loads(json_reports[0].read_text(encoding="utf-8"))
    assert payload["summary"]["not_found"] == 1


def test_strict_contract_match_raises_when_an_endpoint_is_ambiguous(tmp_path):
    contract_path = _build_contract_workbook(
        tmp_path,
        [
            ("Planilha1", "GET", "/v2/pet/{{petId}}"),
            ("Planilha2", "GET", "/v2/pet/{{petId}}"),
        ],
    )
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/:petId"}}]
    )

    use_case = _build_use_case(tmp_path)

    with pytest.raises(StrictContractMatchFailedError) as exc_info:
        use_case.execute_offline(
            contract_file=contract_path, document=document, strict_contract_match=True
        )

    assert "Ambiguous: 1" in str(exc_info.value)


def _build_contract_workbook_with_invalid_field(tmp_path, *, method="GET", uri="/v2/pet/{{petId}}"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Planilha1"
    rows = [
        ["URI", uri],
        ["Método", method],
        ["Resposta caso HTTP Status code 200 - OK"],
        _HEADER_ROW,
        ["1", "dado", "Objeto", None, "SIM"],
        ["1.1", "id", "Alfanumerico", 10, "SIM"],
        ["1.2", "campoRuim", "TipoInvalido", 10, "SIM"],
    ]
    for row in rows:
        sheet.append(row)
    path = tmp_path / "contrato.xlsx"
    workbook.save(path)
    return str(path)


# R2-09A: o ExcelContractValidator agora roda automaticamente (sem flag)
# dentro do use case, e seus diagnósticos são correlacionados com o
# Contract Match Report.


def test_validator_runs_automatically_and_correlates_issues_with_a_matched_entry(tmp_path):
    # Contrato parcialmente inválido (campo com tipo desconhecido) ainda
    # entra no catálogo e ainda casa com a request real — MATCHED e
    # validation_issues coexistem (comportamento real do parser/validator,
    # não inventado).
    contract_path = _build_contract_workbook_with_invalid_field(tmp_path)
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/:petId"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=contract_path, document=document)

    json_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(json_location.path).read_text(encoding="utf-8"))
    matched = next(m for m in payload["matches"] if m["status"] == "MATCHED")
    assert matched["validation_issues"][0]["field"] == "Formato"
    assert matched["validation_issues"][0]["sheet"] == "Planilha1"


def test_validator_issue_on_sheet_without_usable_contract_stays_in_the_general_list(tmp_path):
    workbook_path = tmp_path / "contrato.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "SemMetodoNemUri"
    for row in [["Requisição(Body)"], _HEADER_ROW, ["1", "campo", "TipoInvalido", 10, "SIM"]]:
        sheet.append(row)
    workbook.save(workbook_path)

    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/1"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=str(workbook_path), document=document)

    json_location = next(
        loc for loc in result.artifact_locations if loc.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(json_location.path).read_text(encoding="utf-8"))
    assert payload["validation_issues"][0]["sheet"] == "SemMetodoNemUri"
    for match in payload["matches"]:
        assert "validation_issues" not in match


def test_default_mode_does_not_raise_even_with_unmatched_endpoints(tmp_path):
    # strict_contract_match ausente (default False): comportamento atual
    # preservado — endpoint sem contrato pareado nunca é motivo de falha.
    contract_path = _build_contract_workbook(tmp_path, [("Planilha1", "GET", "/v2/outra-coisa")])
    document = _parse_document(
        [{"name": "Buscar pet", "id": "r1", "request": {"method": "GET", "url": "https://x/v2/pet/1"}}]
    )

    use_case = _build_use_case(tmp_path)
    result = use_case.execute_offline(contract_file=contract_path, document=document)

    assert result.endpoint_outcomes[0].error is None
