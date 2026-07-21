"""Release 2 (Integração de Contratos Excel) — jornada completa ponta a
ponta: contrato declarado numa planilha Excel, pareado por método+path com
uma request real (simulada) de uma Collection Postman, usando o schema
declarado em vez de inferência quando há match — e caindo pra inferência
normalmente quando não há.

Não faz parte da numeração original dos 20 cenários do MVP (SAD original) —
é uma extensão posterior, testada aqui com os mesmos princípios de
aceitação: componentes reais compostos como um wiring de CLI faria (aqui,
GenerateTestsWithContractUseCase, o mesmo use case que bootstrap.py usa
para `generate --contract-file`), só a API do Postman simulada por um
servidor HTTP local.
"""

import json
from pathlib import Path

import openpyxl
from conftest import WORKSPACE_ID, WORKSPACE_NAME, build_app

from api_quality_agent.application.use_cases import GenerateTestsWithContractUseCase
from api_quality_agent.domain.services import (
    ApiAnalysisEngine,
    DiffEngine,
    ManagedBlockMerger,
    SchemaInferenceEngine,
    TestStrategyEngine,
)
from api_quality_agent.generators import PostmanTestGenerator
from api_quality_agent.parsers import ExcelContractParser

_HEADER_ROW = ["Sequencial", "Nome do campo", "Formato", "Tamanho", "Obrigatoriedade", "Regras (Domínio)"]

_COLLECTION_ID = "col-contract"
_COLLECTION_NAME = "Pets API (contrato)"


def _collection_payload() -> dict:
    # Example salvo com status 200: dá a evidência de status que o
    # TestStrategyEngine já exige hoje pra gerar qualquer asserção (mesma
    # exigência de sempre, nada relacionado ao SchemaProvider). A prova de
    # que o schema usado num teste específico é o declarado (não inferido)
    # é o Contract Match Report, verificado nos testes abaixo (MATCHED vs.
    # NOT_FOUND) — não o conteúdo do script em si.
    request: dict = {
        "name": "Buscar pet",
        "id": "req-pet",
        "request": {"method": "GET", "url": "https://api.exemplo.com/pets/:petId"},
        "response": [{"name": "ok", "status": "OK", "code": 200, "header": [], "body": '{"id": 1}'}],
    }
    return {
        "info": {
            "name": _COLLECTION_NAME,
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "item": [request],
    }


def _configure_server(server) -> None:
    server.set_route("/me", status=200, body={"user": {"id": 1, "username": "qa"}})
    server.set_route(
        "/workspaces", status=200, body={"workspaces": [{"id": WORKSPACE_ID, "name": WORKSPACE_NAME}]}
    )
    server.set_route(
        f"/collections?workspace={WORKSPACE_ID}",
        status=200,
        body={"collections": [{"id": _COLLECTION_ID, "uid": _COLLECTION_ID, "name": _COLLECTION_NAME}]},
    )
    server.set_route(
        f"/collections/{_COLLECTION_ID}", status=200, body={"collection": _collection_payload()}
    )


def _write_contract_file(tmp_path) -> str:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pets"
    rows = [
        ["URI", "/pets/{{petId}}"],
        ["Método", "GET"],
        ["Resposta caso HTTP Status code 200 - OK"],
        _HEADER_ROW,
        ["1", "dado", "Objeto", None, "SIM"],
        ["1.1", "id", "Alfanumerico", 10, "SIM"],
        ["1.2", "nome", "Alfanumerico", 50, "SIM"],
    ]
    for row in rows:
        sheet.append(row)
    path = tmp_path / "contrato.xlsx"
    workbook.save(path)
    return str(path)


def _build_generate_with_contract_use_case(app) -> GenerateTestsWithContractUseCase:
    # Mesma composição que bootstrap.py monta para `generate --contract-file`
    # (online) — reaproveitando as dependências já expostas por build_app(),
    # sem alterar conftest.py.
    return GenerateTestsWithContractUseCase(
        ExcelContractParser(),
        ApiAnalysisEngine(),
        SchemaInferenceEngine(),
        TestStrategyEngine(),
        PostmanTestGenerator(),
        ManagedBlockMerger(),
        DiffEngine(),
        app.artifact_repository,
        get_current_workspace_use_case=app.get_current_workspace,
        resolve_collection_use_case=app.resolve_collection,
        collection_repository=app.collection_repository,
    )


def test_generate_with_excel_contract_matches_a_real_collection_request(
    postman_test_server, tmp_path
):
    _configure_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)
    contract_file = _write_contract_file(tmp_path)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=contract_file, collection_id=_COLLECTION_ID)

    outcome = result.endpoint_outcomes[0]
    assert outcome.error is None
    assert outcome.generated_script is not None
    assert "pm.test(" in outcome.generated_script.script

    # A prova de que o endpoint foi pareado com o contrato declarado (não só
    # gerado por inferência, coincidentemente) é o Contract Match Report,
    # salvo automaticamente junto dos demais artefatos.
    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    assert payload["summary"]["matched"] == 1
    assert payload["matches"][0]["status"] == "MATCHED"
    assert payload["matches"][0]["sheet"] == "Pets"
    # Contrato válido: nenhum diagnóstico de validação correlacionado.
    assert "validation_issues" not in payload["matches"][0]
    assert "validation_issues" not in payload

    # O relatório também é salvo em HTML, na mesma pasta de execução.
    html_report = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.html")
    )
    assert Path(html_report.path).exists()


def test_generate_with_excel_contract_falls_back_to_inference_for_unmatched_requests(
    postman_test_server, tmp_path
):
    # Contrato aponta pra um endpoint que não existe na Collection — a
    # request real (com Example salvo) precisa continuar gerando pela
    # inferência de sempre (FallbackSchemaProvider cai pra
    # InferenceSchemaProvider), provando que ausência de match não quebra
    # nem esvazia o que já funcionava antes desta funcionalidade existir.
    _configure_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "OutroEndpoint"
    for row in [["URI", "/nao-existe"], ["Método", "GET"]]:
        sheet.append(row)
    contract_file = tmp_path / "contrato_sem_match.xlsx"
    workbook.save(contract_file)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=str(contract_file), collection_id=_COLLECTION_ID)

    outcome = result.endpoint_outcomes[0]
    assert outcome.error is None
    # A request TEM Example salvo — sem contrato pareado, o fallback pra
    # inferência precisa gerar a asserção normalmente, como sempre gerou.
    assert "pm.test(" in outcome.generated_script.script
    assert "pm.response.to.have.status(200)" in outcome.generated_script.script

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    assert payload["summary"]["not_found"] == 1


_GATEWAY_COLLECTION_ID = "col-contract-gateway"
_GATEWAY_COLLECTION_NAME = "Pets API (gateway)"


def _gateway_collection_payload() -> dict:
    # Collection representativa de um cenário real de gateway: a request
    # carrega um prefixo fixo ("/api") que o contrato declarado não tem —
    # exatamente o problema que --collection-path-prefix resolve.
    request: dict = {
        "name": "Buscar pet",
        "id": "req-pet-gateway",
        "request": {"method": "GET", "url": "https://api.exemplo.com/api/pets/:petId"},
        "response": [{"name": "ok", "status": "OK", "code": 200, "header": [], "body": '{"id": 1}'}],
    }
    return {
        "info": {
            "name": _GATEWAY_COLLECTION_NAME,
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "item": [request],
    }


def _configure_gateway_server(server) -> None:
    server.set_route("/me", status=200, body={"user": {"id": 1, "username": "qa"}})
    server.set_route(
        "/workspaces", status=200, body={"workspaces": [{"id": WORKSPACE_ID, "name": WORKSPACE_NAME}]}
    )
    server.set_route(
        f"/collections?workspace={WORKSPACE_ID}",
        status=200,
        body={
            "collections": [
                {"id": _GATEWAY_COLLECTION_ID, "uid": _GATEWAY_COLLECTION_ID, "name": _GATEWAY_COLLECTION_NAME}
            ]
        },
    )
    server.set_route(
        f"/collections/{_GATEWAY_COLLECTION_ID}",
        status=200,
        body={"collection": _gateway_collection_payload()},
    )


def test_generate_without_collection_path_prefix_does_not_match_a_gateway_collection(
    postman_test_server, tmp_path
):
    _configure_gateway_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)
    contract_file = _write_contract_file(tmp_path)  # contrato declara "/pets/{{petId}}", sem "/api"

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=contract_file, collection_id=_GATEWAY_COLLECTION_ID)

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    assert payload["summary"]["not_found"] == 1


def test_generate_with_collection_path_prefix_matches_a_gateway_collection(
    postman_test_server, tmp_path
):
    _configure_gateway_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)
    contract_file = _write_contract_file(tmp_path)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(
        contract_file=contract_file,
        collection_id=_GATEWAY_COLLECTION_ID,
        collection_path_prefix="/api",
    )

    outcome = result.endpoint_outcomes[0]
    assert outcome.error is None
    assert "pm.test(" in outcome.generated_script.script

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    assert payload["summary"]["matched"] == 1
    assert payload["matches"][0]["status"] == "MATCHED"
    assert payload["matches"][0]["sheet"] == "Pets"


# R2-09A: correlação entre os diagnósticos do ExcelContractValidator (agora
# rodando automaticamente, sem flag) e o Contract Match Report. Também
# confirma regressão zero de --collection-path-prefix e --strict-contract-match.


def test_generate_with_partially_invalid_contract_still_matches_and_shows_the_diagnostic(
    postman_test_server, tmp_path
):
    # Contrato com um campo de schema inválido ainda entra no catálogo e
    # ainda casa com a request real (parser e validator são passes
    # desacoplados) — MATCHED e validation_issues coexistem no relatório.
    _configure_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pets"
    rows = [
        ["URI", "/pets/{{petId}}"],
        ["Método", "GET"],
        ["Resposta caso HTTP Status code 200 - OK"],
        _HEADER_ROW,
        ["1", "dado", "Objeto", None, "SIM"],
        ["1.1", "id", "Alfanumerico", 10, "SIM"],
        ["1.2", "campoRuim", "TipoInvalido", 10, "SIM"],
    ]
    for row in rows:
        sheet.append(row)
    contract_file = tmp_path / "contrato_parcial.xlsx"
    workbook.save(contract_file)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=str(contract_file), collection_id=_COLLECTION_ID)

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    matched_entry = payload["matches"][0]
    assert matched_entry["status"] == "MATCHED"
    assert matched_entry["validation_issues"][0]["field"] == "Formato"
    assert matched_entry["validation_issues"][0]["sheet"] == "Pets"


def test_generate_with_invalid_sheet_keeps_its_diagnostic_only_in_the_general_list(
    postman_test_server, tmp_path
):
    # Aba sem Método/URI nunca produz contrato utilizável — o diagnóstico
    # correspondente nunca deve ser forçado pra dentro de matches[].
    _configure_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pets"
    for row in [["URI", "/pets/{{petId}}"], ["Método", "GET"]]:
        sheet.append(row)
    invalid_sheet = workbook.create_sheet("SemMetodoNemUri")
    for row in [["Requisição(Body)"], _HEADER_ROW, ["1", "campo", "TipoInvalido", 10, "SIM"]]:
        invalid_sheet.append(row)
    contract_file = tmp_path / "contrato_com_aba_invalida.xlsx"
    workbook.save(contract_file)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=str(contract_file), collection_id=_COLLECTION_ID)

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    assert any(issue["sheet"] == "SemMetodoNemUri" for issue in payload["validation_issues"])
    for match in payload["matches"]:
        if match["status"] != "MATCHED":
            continue
        assert all(
            issue["sheet"] != "SemMetodoNemUri" for issue in match.get("validation_issues", [])
        )


def test_generate_ambiguous_match_carries_different_issues_per_candidate(
    postman_test_server, tmp_path
):
    _configure_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)

    workbook = openpyxl.Workbook()
    sheet_a = workbook.active
    sheet_a.title = "PetsA"
    for row in [
        ["URI", "/pets/{{petId}}"],
        ["Método", "GET"],
        ["Resposta caso HTTP Status code 200 - OK"],
        _HEADER_ROW,
        ["1", "campoRuimA", "TipoInvalido", 10, "SIM"],
    ]:
        sheet_a.append(row)
    sheet_b = workbook.create_sheet("PetsB")
    for row in [
        ["URI", "/pets/{{petId}}"],
        ["Método", "GET"],
        ["Resposta caso HTTP Status code 200 - OK"],
        _HEADER_ROW,
        ["1", "campoRuimB", "OutroTipoInvalido", 10, "SIM"],
    ]:
        sheet_b.append(row)
    contract_file = tmp_path / "contrato_ambiguo.xlsx"
    workbook.save(contract_file)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=str(contract_file), collection_id=_COLLECTION_ID)

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    ambiguous_entry = next(m for m in payload["matches"] if m["status"] == "AMBIGUOUS")
    by_sheet = {c["sheet"]: c["issues"] for c in ambiguous_entry["candidate_validation_issues"]}
    assert by_sheet["PetsA"][0]["field"] == "Formato"
    assert by_sheet["PetsA"][0]["message"] != by_sheet["PetsB"][0]["message"]


def test_generate_not_found_never_carries_correlation_even_with_a_similarly_named_sheet(
    postman_test_server, tmp_path
):
    _configure_server(postman_test_server)
    app = build_app(postman_test_server, tmp_path)
    app.select_workspace.execute(workspace_id=WORKSPACE_ID)

    # Contrato aponta pra um endpoint inexistente na Collection (NOT_FOUND),
    # mas o nome da aba parece muito com a Collection real — mesmo assim,
    # nenhuma correlação deve ocorrer.
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PetsMuitoParecidoComOReal"
    for row in [
        ["URI", "/nao-existe"],
        ["Método", "GET"],
        ["Resposta caso HTTP Status code 200 - OK"],
        _HEADER_ROW,
        ["1", "campoRuim", "TipoInvalido", 10, "SIM"],
    ]:
        sheet.append(row)
    contract_file = tmp_path / "contrato_not_found.xlsx"
    workbook.save(contract_file)

    use_case = _build_generate_with_contract_use_case(app)
    result = use_case.execute_online(contract_file=str(contract_file), collection_id=_COLLECTION_ID)

    match_json = next(
        location for location in result.artifact_locations if location.path.endswith("contract-match-report.json")
    )
    payload = json.loads(Path(match_json.path).read_text(encoding="utf-8"))
    not_found_entry = next(m for m in payload["matches"] if m["status"] == "NOT_FOUND")
    assert "validation_issues" not in not_found_entry
    assert "candidate_validation_issues" not in not_found_entry
    # O diagnóstico da aba continua disponível na lista geral.
    assert any(issue["sheet"] == "PetsMuitoParecidoComOReal" for issue in payload["validation_issues"])
